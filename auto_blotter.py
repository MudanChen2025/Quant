# This script will read toms blotter txt and generate a full version with fee and then send a email to Joey

import pandas as pd
import numpy as np
import csv
from datetime import datetime, timedelta
from collections import deque
from blpapi import Session, SessionOptions
from blp import blp
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

def traditional_round(num, n):
    multiplier = 10 ** n
    return int(num * multiplier + 0.5) / multiplier

def read_txt_to_dataframe(file_path):
    data = []
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        # header start with 'Trader'
        header_index = next(i for i, line in enumerate(lines) if line.startswith('Trader'))
        headers =['Trader','Counterpar','B/S','Long Description','ISIN','Amount','Trd Dt','As of Dt','Transac','Trm Date','Stl Date','Coupon','Crcy','Price','Principal','Accr Int','Settlement Amount','Repo Rte Haircut','Unadj Term Money','Tkt #','Execution Order Identifier','Broker Commissi','Stamp Duty Amou','Transaction Lev','Exchange Fee Am','Miscellaneous F','Commission']
        # get the index of each item in headers
        headers_index = [lines[header_index].index(header) for header in headers]
        headers_index.append(len(lines[header_index]))
        for line in lines:
            if line.startswith('-') or line.startswith(' ') or line.startswith('START OF REPORT') or line.startswith('END OF REPORT') or line.startswith('Trader'):
                continue
            # split the line by index in headers_index
            values = [line[headers_index[i]:headers_index[i+1]].strip() for i in range(len(headers_index)-1)]
            data.append(values)
    
    # create dataframe and do basic cleaning
    df = pd.DataFrame(data, columns=headers)
    df = df.replace('', np.nan)
    df = df.dropna(axis=0,how='all')
    df.reset_index(drop=True, inplace=True)
    df['Amount'] = pd.to_numeric(df['Amount'].str.replace(',', ''))
    df['Price'] = pd.to_numeric(df['Price'])
    df['Long Description'] = pd.to_numeric(df['Long Description'])
    # don't rely on the existing Principal column, recalculate it
    df['Principal'] = df['Amount']*df['Price']
    #drop columns
    df.drop(columns=['Broker Commissi','Stamp Duty Amou','Transaction Lev','Exchange Fee Am','Miscellaneous F','Commission'], inplace=True)
    # if Execution Order Identifier is NA, fill it with a random number, don't repeat
    for i in range(len(df)):
        if pd.isna(df.loc[i, 'Execution Order Identifier']):
            df.loc[i, 'Execution Order Identifier'] = hash(str(i))
    return df

# read in data sources
current_date = datetime.now().strftime("%Y%m%d")
file_path = "C:\\Users\\jennychen\\Desktop\\pnl_automation_project\\daily_blotter\\input\\toms_trade_blotter_{}.txt".format(current_date) 
df = read_txt_to_dataframe(file_path)
df_fee = pd.read_excel('C:\\Users\\jennychen\\Desktop\\pnl_automation_project\\daily_blotter\\input\\SG Fee and Comm.xlsx', sheet_name="Fee")
df_commission = pd.read_excel('C:\\Users\\jennychen\\Desktop\\pnl_automation_project\\daily_blotter\\input\\SG Fee and Comm.xlsx', sheet_name="Commission")

# groupby blotter to mather order based on Execution Order Identifier
df_mother = df.groupby(['Trader','Counterpar','B/S','Long Description','ISIN','Trd Dt','As of Dt','Transac','Stl Date','Crcy','Execution Order Identifier']).agg({'Principal':'sum','Amount':'sum'}).reset_index()
df_mother["Average Price"] = (df_mother['Principal']/df_mother['Amount']).apply(lambda x: traditional_round(x, 4))
df_mother["Gross Amount"] = df_mother['Amount'] * df_mother['Average Price']
df_mother.loc[df_mother['Crcy'] == 'JP', 'Gross Amount'] = df_mother[df_mother['Crcy'] == 'JP']['Gross Amount'].apply(lambda x: traditional_round(x, 0))
# calculate commission
df_mother = df_mother.merge(df_commission,left_on=['Counterpar','Crcy'], right_on=['Counterparty','CURRENCY'],how='left')
df_mother = df_mother.rename(columns={'Counterparty':'CM_Counterparty','CURRENCY':'CM_Currency','FEE_CHARGE':'CM_fee','Min':'CM_min'})
df_mother['CM_fee'] = df_mother['CM_fee'].fillna(0)
df_mother['Commission'] = df_mother['CM_fee'] * df_mother['Gross Amount']
df_mother.loc[df_mother['Commission'] < df_mother['CM_min'], 'Commission'] = df_mother['CM_min']
df_mother.drop(columns=['CM_Counterparty','CM_Currency','CM_fee','CM_min'], inplace=True)
df_mother['Commission'] = np.where(df_mother['Crcy'] == 'JPY', 
                                   df_mother['Commission'].apply(lambda x: traditional_round(x, 0)), 
                                   df_mother['Commission'].apply(lambda x: traditional_round(x, 2)))
# calculate all fee
df_mother['Sec Fee'] = 0
df_mother['Stamp Duty'] = 0
df_mother['Transaction Levy'] = 0
df_mother['Trading Fee'] = 0
df_mother['AFRC Transaction Levy'] = 0
for i, row in df_mother.iterrows():
    currency = row['Crcy']
    operation = row['B/S']
    principle = row['Gross Amount']
    df_fee_schema = df_fee[(df_fee['Currency'] == currency) & (df_fee['Operations'].str[:1] == operation)]
    for index, fee_row in df_fee_schema.iterrows():
        charge_name = fee_row['Charges Name']
        charge_percentage = fee_row['Value']
        charge_round_method = fee_row['Round Method']
        charge_decimal = fee_row['Decimal']
        charge_min = fee_row['Maximum Charges']
        cost = principle * charge_percentage * 0.01
        if charge_round_method == 'Round Up':
            multiplier = 10 ** charge_decimal
            cost = math.ceil(cost * multiplier) / multiplier
        else:
            cost = traditional_round(cost, 2)
        if cost < charge_min:
            cost = charge_min
        df_mother.loc[i, charge_name] = cost

# calculate final settlement amount
def calculate_settlement_amount(row):
    if row['B/S'] == "B":
        return row['Gross Amount'] + row['Commission'] + row['Sec Fee'] + row['Stamp Duty'] + row['Transaction Levy'] + row['Trading Fee'] + row['AFRC Transaction Levy']
    else:
        return row['Gross Amount'] - row['Commission'] - row['Sec Fee'] - row['Stamp Duty'] - row['Transaction Levy'] - row['Trading Fee'] - row['AFRC Transaction Levy']
df_mother['Settlement Amount'] = df_mother.apply(calculate_settlement_amount, axis=1)

df_mother.rename(columns={'Counterpar':'Counterparty','Long Description':'Security Description','Trd Dt':'Trade Date','Transac':'Transaction Date','Stl Date':'Settle Date','Crcy':'Currency'},inplace=True)

trade_date = datetime.strptime(df_mother['Trade Date'][0], "%m/%d/%y").strftime('%Y%m%d')
file_name = "C:\\Users\\jennychen\\Desktop\\pnl_automation_project\\daily_blotter\\output\\toms_trade_blotter_{}.xlsx".format(trade_date) 

# save result
wb = Workbook()
ws = wb.active
for r_idx, row in enumerate(dataframe_to_rows(df_mother, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws[1]:
    cell.fill = header_fill
    
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(file_name)

#generate email
smtp_server = 'outlook.office365.com'
sender_email = 'jennychen@tfisec.com'
sender_password = 'Cmd1998122!'
receiver_email = 'jennychen@tfisec.com'
message = MIMEMultipart()
message['From'] = sender_email
message['To'] = receiver_email
message['Subject'] = 'Toms Trade Blotter on ' + f"{trade_date}"
body = 'Please see the attachment for toms trade blotter on ' + f"{trade_date}." + "\n\n"
message.attach(MIMEText(body, 'plain'))
attachment_path = file_name
attachment = open(attachment_path, 'rb')
part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
file_name_ = f"toms_trade_blotter_{trade_date}.xlsx" 
part.add_header('Content-Disposition', f"attachment; filename= {file_name_}")
message.attach(part)
server = smtplib.SMTP(smtp_server, 587)
server.starttls()
server.login(sender_email, sender_password)
text = message.as_string()
server.sendmail(sender_email, receiver_email, text)
server.quit()
attachment.close()